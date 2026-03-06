#!/usr/bin/env python3
"""
SuppTree Excel → JSON Converter
Reads SuppTree_KOMPLETT_FINAL_v2.xlsx and generates JSON data files for the app.
"""

import json
import re
import sys
import io
import os
from collections import defaultdict

# Fix encoding for Windows console
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

import openpyxl

# === CONFIG ===
EXCEL_PATH = r"C:\Users\feich\OneDrive\Desktop\Recherche Wissendatenbank\Supplements Excel\Zusammengeführte Tabellen\KOMPLETT\SuppTree_KOMPLETT_FINAL_v2.xlsx"
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "www", "data")

# === HELPERS ===

def safe_str(val):
    """Convert cell value to string, handling None."""
    if val is None:
        return ""
    return str(val).strip()

def safe_float(val, default=0):
    """Convert cell value to float."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default

def safe_int(val, default=0):
    """Convert cell value to int."""
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default

def safe_bool(val):
    """Convert cell value to boolean."""
    if val is None:
        return False
    s = str(val).strip().lower()
    return s in ('true', 'ja', 'yes', '1', 'wahr')

def make_id(name):
    """Create a URL-safe ID from a name."""
    if not name:
        return ""
    s = name.lower().strip()
    # Replace umlauts
    s = s.replace('ä', 'ae').replace('ö', 'oe').replace('ü', 'ue').replace('ß', 'ss')
    s = s.replace('Ä', 'ae').replace('Ö', 'oe').replace('Ü', 'ue')
    # Replace special chars with hyphens
    s = re.sub(r'[^a-z0-9]+', '-', s)
    s = re.sub(r'-+', '-', s).strip('-')
    return s

def split_list(val):
    """Split a comma/semicolon separated string into a list."""
    if not val:
        return []
    s = safe_str(val)
    if not s:
        return []
    # Split on comma or semicolon
    parts = re.split(r'[,;]+', s)
    return [p.strip() for p in parts if p.strip()]

def read_sheet(wb, sheet_name):
    """Read a sheet into a list of dicts using first row as headers."""
    if sheet_name not in wb.sheetnames:
        print(f"  WARNING: Sheet '{sheet_name}' not found!")
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [safe_str(h).strip() for h in rows[0]]
    data = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        entry = {}
        for i, h in enumerate(headers):
            if h and i < len(row):
                entry[h] = row[i]
        data.append(entry)
    return data

def risiko_to_ampel(risiko_level):
    """Map risiko_level string to ampel color."""
    if not risiko_level:
        return "gruen"
    s = safe_str(risiko_level).lower()
    if 'hoch' in s or 'kontraindiziert' in s or 'gefährlich' in s or 'gefaehrlich' in s:
        return "rot"
    if 'mittel' in s or 'moderat' in s:
        return "gelb"
    if 'niedrig' in s or 'gering' in s or 'positiv' in s or 'neutral' in s:
        return "gruen"
    # Default based on keywords
    if 'vorsicht' in s or 'achtung' in s:
        return "gelb"
    return "gruen"

# === NORMALIZATION FOR MEDICATION GROUPS ===

# Mapping of common variations to canonical group names
GROUP_NORMALIZATION = {
    # ACE-Hemmer variants
    'ace-hemmer': 'ACE-Hemmer',
    'ace hemmer': 'ACE-Hemmer',
    'ace-hemmer / arbs': 'ACE-Hemmer / ARBs',
    'ace-hemmer/arbs': 'ACE-Hemmer / ARBs',
    'ace-hemmer/arbs/k-sparende diuretika': 'ACE-Hemmer / ARBs',
    'angiotensin-rezeptorblocker': 'ACE-Hemmer / ARBs',
    'arbs': 'ACE-Hemmer / ARBs',
    # Antikoagulanzien
    'antikoagulanzien': 'Antikoagulanzien (Blutverdünner)',
    'blutverdünner': 'Antikoagulanzien (Blutverdünner)',
    'blutverduenner': 'Antikoagulanzien (Blutverdünner)',
    'antikoagulanzien (vitamin-k-antagonisten)': 'Antikoagulanzien (Blutverdünner)',
    'vitamin-k-antagonisten': 'Antikoagulanzien (Blutverdünner)',
    'direkte orale antikoagulanzien': 'Antikoagulanzien (Blutverdünner)',
    'doaks': 'Antikoagulanzien (Blutverdünner)',
    'noak': 'Antikoagulanzien (Blutverdünner)',
    # Antidepressiva
    'antidepressiva': 'Antidepressiva',
    'ssri': 'Antidepressiva (SSRI)',
    'ssris': 'Antidepressiva (SSRI)',
    'snri': 'Antidepressiva (SNRI)',
    'snris': 'Antidepressiva (SNRI)',
    'trizyklische antidepressiva': 'Antidepressiva (Trizyklische)',
    'maoi': 'Antidepressiva (MAO-Hemmer)',
    'mao-hemmer': 'Antidepressiva (MAO-Hemmer)',
    # Diabetes
    'antidiabetika': 'Antidiabetika',
    'diabetes-medikamente': 'Antidiabetika',
    'metformin': 'Antidiabetika (Metformin)',
    'insulin': 'Antidiabetika (Insulin)',
    'sulfonylharnstoffe': 'Antidiabetika (Sulfonylharnstoffe)',
    # Statine
    'statine': 'Statine (Cholesterinsenker)',
    'cholesterinsenker': 'Statine (Cholesterinsenker)',
    'hmg-coa-reduktase-hemmer': 'Statine (Cholesterinsenker)',
    # Schilddruese
    'schilddrüsenmedikamente': 'Schilddrüsenmedikamente',
    'schilddruesenmedikamente': 'Schilddrüsenmedikamente',
    'levothyroxin': 'Schilddrüsenmedikamente',
    'l-thyroxin': 'Schilddrüsenmedikamente',
    # PPI
    'protonenpumpenhemmer': 'Protonenpumpenhemmer (PPI)',
    'ppi': 'Protonenpumpenhemmer (PPI)',
    'protonenpumpeninhibitoren': 'Protonenpumpenhemmer (PPI)',
    'säureblocker': 'Protonenpumpenhemmer (PPI)',
    # Antibiotika
    'antibiotika': 'Antibiotika',
    'fluorchinolone': 'Antibiotika (Fluorchinolone)',
    'tetracycline': 'Antibiotika (Tetracycline)',
    'aminoglykoside': 'Antibiotika (Aminoglykoside)',
    # NSAIDs
    'nsaids': 'NSAIDs (Schmerzmittel)',
    'nsar': 'NSAIDs (Schmerzmittel)',
    'schmerzmittel': 'NSAIDs (Schmerzmittel)',
    'nicht-steroidale antirheumatika': 'NSAIDs (Schmerzmittel)',
    # Immunsuppressiva
    'immunsuppressiva': 'Immunsuppressiva',
    'immunmodulatoren': 'Immunsuppressiva',
    # Diuretika
    'diuretika': 'Diuretika',
    'thiaziddiuretika': 'Diuretika (Thiazide)',
    'thiazide': 'Diuretika (Thiazide)',
    'schleifendiuretika': 'Diuretika (Schleifendiuretika)',
    'kaliumsparende diuretika': 'Diuretika (Kaliumsparend)',
    'k-sparende diuretika': 'Diuretika (Kaliumsparend)',
    # Hormone
    'hormonelle verhütungsmittel': 'Hormonelle Verhütungsmittel',
    'antibabypille': 'Hormonelle Verhütungsmittel',
    'kontrazeptiva': 'Hormonelle Verhütungsmittel',
    'hormonersatztherapie': 'Hormonersatztherapie (HRT)',
    'hrt': 'Hormonersatztherapie (HRT)',
    # Corticosteroide
    'corticosteroide': 'Corticosteroide',
    'kortikosteroide': 'Corticosteroide',
    'glukokortikoide': 'Corticosteroide',
    'kortison': 'Corticosteroide',
    # Antiepileptika
    'antiepileptika': 'Antiepileptika',
    'antikonvulsiva': 'Antiepileptika',
    # Zytostatika / Chemo
    'zytostatika': 'Zytostatika (Chemotherapie)',
    'chemotherapeutika': 'Zytostatika (Chemotherapie)',
    'chemotherapie': 'Zytostatika (Chemotherapie)',
    'antineoplastika': 'Zytostatika (Chemotherapie)',
    # Betablocker
    'betablocker': 'Betablocker',
    'beta-blocker': 'Betablocker',
    'beta-adrenozeptor-antagonisten': 'Betablocker',
    # Calciumkanalblocker
    'calciumkanalblocker': 'Calciumkanalblocker',
    'kalziumantagonisten': 'Calciumkanalblocker',
    'calcium-antagonisten': 'Calciumkanalblocker',
    # Lithium
    'lithium': 'Lithium',
    # Bisphosphonate
    'bisphosphonate': 'Bisphosphonate',
    # Antazida
    'antazida': 'Antazida',
    # Herzglykoside
    'herzglykoside': 'Herzglykoside',
    'digitalis': 'Herzglykoside',
    'digoxin': 'Herzglykoside',
    # Sedativa / Benzodiazepine
    'benzodiazepine': 'Sedativa / Benzodiazepine',
    'sedativa': 'Sedativa / Benzodiazepine',
    'schlafmittel': 'Sedativa / Benzodiazepine',
    # Antipsychotika
    'antipsychotika': 'Antipsychotika',
    'neuroleptika': 'Antipsychotika',
    # HIV
    'antiretrovirale medikamente': 'Antiretrovirale Medikamente (HIV)',
    'hiv-medikamente': 'Antiretrovirale Medikamente (HIV)',
}

def normalize_group_name(raw_name):
    """Normalize a medication group name to a canonical form."""
    if not raw_name:
        return ""
    s = safe_str(raw_name).strip()
    key = s.lower().strip()
    # Remove trailing parenthetical notes
    key = re.sub(r'\s*\(z\.?b\.?[^)]*\)', '', key)
    key = key.strip()

    if key in GROUP_NORMALIZATION:
        return GROUP_NORMALIZATION[key]

    # Try partial matches
    for pattern, canonical in GROUP_NORMALIZATION.items():
        if pattern in key or key in pattern:
            return canonical

    # Return original with first letter capitalized
    return s

def extract_med_names(beispiele_str):
    """Extract individual medication names from a 'beispiele' string."""
    if not beispiele_str:
        return []
    s = safe_str(beispiele_str)
    # Handle "GroupName (Med1, Med2, Med3)" pattern:
    # Extract content inside parentheses as the actual medication names
    paren_match = re.match(r'^[^(]+\(([^)]+)\)(.*)$', s)
    if paren_match:
        inner = paren_match.group(1)
        after = paren_match.group(2).strip().strip(',;')
        s = inner + (',' + after if after else '')
    # Split on comma, semicolon, slash, "und", "oder"
    parts = re.split(r'[,;/]+|\bund\b|\boder\b', s)
    names = []
    for p in parts:
        p = p.strip()
        # Remove content in closed parentheses (generic names etc.)
        p = re.sub(r'\([^)]*\)', '', p).strip()
        # Remove unclosed parentheses (leftover fragments)
        p = re.sub(r'\(.*$', '', p).strip()
        p = re.sub(r'^[^)]*\)', '', p).strip()
        # Remove dosage info
        p = re.sub(r'\d+\s*mg\b', '', p).strip()
        # Remove leading/trailing special chars and emoji prefixes
        p = p.strip('- .')
        p = re.sub(r'^[\u2600-\u27BF\u2B50\u26A0\u274C\u2705\u26D4\uFE0F⚠️⛔❌✅]+\s*', '', p).strip()
        if p and len(p) > 1:
            # Filter out junk entries that are descriptions, not real medication names
            pl = p.lower()
            if pl.startswith('alle ') or pl.startswith('generell') or pl == 'etc':
                continue
            if pl.endswith(' etc') or pl.endswith(' etc.'):
                continue
            if pl.startswith('z.b.') or pl.startswith('z. b.'):
                continue
            if pl.startswith('theoretisch ') or pl.startswith('alle formen:'):
                continue
            if '-haltige ' in pl or '-praeparate' in pl or '-präparate' in pl:
                continue
            if pl.endswith('-generika') or pl.endswith(' generika'):
                continue
            if re.match(r'^(alle|jede|sämtliche|saemtliche)\b', pl, re.IGNORECASE):
                continue
            names.append(p)
    return names

# === FORM KEYWORDS FOR AUTO-DETECTION ===
FORM_KEYWORDS = {
    'Spritze': ['spritze', 'injektion', 'injekt', 'subkutan', 'i.v.', 'i.m.'],
    'Tropfen': ['tropfen', 'drops'],
    'Salbe': ['salbe', 'creme', 'gel', 'topisch', 'extern'],
    'Spray': ['spray', 'nasenspray', 'inhalat'],
    'Kapsel': ['kapsel', 'softgel'],
    'Sirup': ['sirup', 'saft', 'lösung', 'loesung'],
    'Pflaster': ['pflaster', 'patch', 'transdermal'],
    'Suppositorium': ['suppositor', 'zäpfchen', 'zaepfchen', 'rektal'],
    'Pulver': ['pulver', 'granulat'],
}

def guess_form(name, gruppe):
    """Guess the dosage form from medication/group name."""
    combined = f"{name} {gruppe}".lower()
    for form, keywords in FORM_KEYWORDS.items():
        for kw in keywords:
            if kw in combined:
                return form
    return "Tablette"


# ==========================================================
# MAIN EXTRACTION
# ==========================================================

def main():
    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # -------------------------------------------------------
    # 1. SUPPLEMENTS (wirkstoffe.json)
    # -------------------------------------------------------
    print("\n=== 1. Supplements (wirkstoffe.json) ===")
    supps = read_sheet(wb, 'Supplements_Basis')
    print(f"  Raw rows: {len(supps)}")

    wirkstoffe = []
    supp_id_map = {}  # name -> id mapping for cross-references

    for row in supps:
        sid = safe_str(row.get('id', ''))
        name = safe_str(row.get('name', ''))
        if not name:
            continue
        if not sid:
            sid = make_id(name)

        # Build entry
        entry = {
            "id": sid,
            "name": name,
            "name_en": safe_str(row.get('name_EN', '')),
            "auch_bekannt_als": safe_str(row.get('auch_bekannt_als', '')),
            "kategorie": safe_str(row.get('kategorie', '')),
            "unterkategorie": safe_str(row.get('unterkategorie', '')),
            "wirkung": safe_str(row.get('wirkung', '')),
            "wirkung_kurz": safe_str(row.get('wirkung_kurz', '')),
            "einnahme_zeitpunkt": safe_str(row.get('einnahme_zeitpunkt', '')),
            "einnahme_mit": safe_str(row.get('einnahme_mit', '')),
            "einnahme_ohne": safe_str(row.get('einnahme_ohne', '')),
            "einnahme_hinweis": safe_str(row.get('einnahme_hinweis', '')),
            "fettloeslich": safe_bool(row.get('fettloeslich', False)),
            "wasserloeslich": safe_bool(row.get('wasserloeslich', False)),
            "kombiniert_gut_mit": split_list(row.get('kombiniert_gut_mit', '')),
            "nicht_zusammen_mit": split_list(row.get('nicht_zusammen_mit', '')),
            "zeitlicher_abstand_h": safe_float(row.get('zeitlicher_abstand_h', 0)),
            "warnhinweise": safe_str(row.get('warnhinweise', '')),
            "kontraindikationen": safe_str(row.get('kontraindikationen', '')),
            "max_einzeldosis": safe_str(row.get('max_einzeldosis', '')),
            "vegan_verfuegbar": safe_bool(row.get('vegan_verfuegbar', False)),
        }

        wirkstoffe.append(entry)
        # Map various names to this ID
        supp_id_map[name.lower()] = sid
        supp_id_map[sid] = sid
        for alias in split_list(row.get('auch_bekannt_als', '')):
            supp_id_map[alias.lower()] = sid

    print(f"  Exported: {len(wirkstoffe)} wirkstoffe")

    with open(os.path.join(OUTPUT_DIR, 'wirkstoffe.json'), 'w', encoding='utf-8') as f:
        json.dump(wirkstoffe, f, ensure_ascii=False, indent=2)
    print(f"  Written: wirkstoffe.json")

    # -------------------------------------------------------
    # 2. SUPPLEMENT INTERACTIONS (supplement_interactions.json)
    # -------------------------------------------------------
    print("\n=== 2. Supplement Interactions ===")
    ww = read_sheet(wb, 'Wechselwirkungen_Supplements')
    print(f"  Raw rows: {len(ww)}")

    supp_interactions = []
    for row in ww:
        a = safe_str(row.get('supplement_a', ''))
        b = safe_str(row.get('supplement_b', ''))
        if not a or not b:
            continue

        # Resolve IDs
        a_id = supp_id_map.get(a.lower(), make_id(a))
        b_id = supp_id_map.get(b.lower(), make_id(b))

        entry = {
            "supplement_a": a_id,
            "supplement_a_name": a,
            "supplement_b": b_id,
            "supplement_b_name": b,
            "typ": safe_str(row.get('typ', 'neutral')),
            "staerke": safe_str(row.get('staerke', '')),
            "richtung": safe_str(row.get('richtung', 'bidirektional')),
            "mechanismus": safe_str(row.get('mechanismus', '')),
            "praktische_auswirkung": safe_str(row.get('praktische_auswirkung', '')),
            "empfehlung": safe_str(row.get('empfehlung', '')),
            "zeitabstand_h": safe_float(row.get('zeitabstand_h', 0)),
            "evidenz": safe_str(row.get('evidenz', '')),
        }
        supp_interactions.append(entry)

    print(f"  Exported: {len(supp_interactions)} supplement interactions")

    with open(os.path.join(OUTPUT_DIR, 'supplement_interactions.json'), 'w', encoding='utf-8') as f:
        json.dump(supp_interactions, f, ensure_ascii=False, indent=2)
    print(f"  Written: supplement_interactions.json")

    # -------------------------------------------------------
    # 3. MEDICATION INTERACTIONS (med_supp_interactions.json)
    # -------------------------------------------------------
    print("\n=== 3. Medication-Supplement Interactions ===")
    med_int = read_sheet(wb, 'Medikamenten_Interaktionen')
    print(f"  Raw rows: {len(med_int)}")

    med_supp_interactions = []
    all_raw_groups = set()
    group_medications = defaultdict(set)  # group_name -> set of medication names

    for row in med_int:
        supp_id = safe_str(row.get('supplement_id', ''))
        supp_name = safe_str(row.get('supplement_name', ''))
        med_gruppe = safe_str(row.get('medikament_gruppe', ''))
        med_beispiele = safe_str(row.get('medikament_beispiele', ''))
        risiko = safe_str(row.get('risiko_level', ''))

        if not supp_name or not med_gruppe:
            continue

        # Resolve supplement ID
        if not supp_id:
            supp_id = supp_id_map.get(supp_name.lower(), make_id(supp_name))

        # Normalize group
        norm_gruppe = normalize_group_name(med_gruppe)
        all_raw_groups.add(med_gruppe)

        # Collect medication names for this group
        med_names = extract_med_names(med_beispiele)
        for mn in med_names:
            group_medications[norm_gruppe].add(mn)

        entry = {
            "supplement_id": supp_id,
            "supplement_name": supp_name,
            "medikament_gruppe": norm_gruppe,
            "medikament_gruppe_raw": med_gruppe,
            "medikament_beispiele": med_beispiele,
            "interaktions_typ": safe_str(row.get('interaktions_typ', '')),
            "mechanismus": safe_str(row.get('mechanismus', '')),
            "risiko_level": risiko,
            "klinische_relevanz": safe_str(row.get('klinische_relevanz', '')),
            "empfehlung": safe_str(row.get('empfehlung', '')),
            "ampel": risiko_to_ampel(risiko),
        }
        med_supp_interactions.append(entry)

    print(f"  Exported: {len(med_supp_interactions)} med-supp interactions")
    print(f"  Raw unique groups: {len(all_raw_groups)}")

    with open(os.path.join(OUTPUT_DIR, 'med_supp_interactions.json'), 'w', encoding='utf-8') as f:
        json.dump(med_supp_interactions, f, ensure_ascii=False, indent=2)
    print(f"  Written: med_supp_interactions.json")

    # -------------------------------------------------------
    # 4. MEDICATION DATABASE (medikamente_db.json + medikament_gruppen.json)
    # -------------------------------------------------------
    print("\n=== 4. Medication Database ===")

    # Build normalized groups with their medications
    gruppen_list = []
    alle_medikamente = []

    # Sort groups alphabetically
    sorted_groups = sorted(group_medications.keys())

    for idx, grp_name in enumerate(sorted_groups, 1):
        grp_id = f"grp_{idx:04d}"
        meds = sorted(group_medications[grp_name])

        # Build normalized search terms
        namen_norm = [grp_name.lower()]
        # Add parts without parentheses
        clean = re.sub(r'\([^)]*\)', '', grp_name).strip()
        if clean.lower() != grp_name.lower():
            namen_norm.append(clean.lower())
        # Add individual words
        words = re.split(r'[\s/\-]+', grp_name.lower())
        for w in words:
            if len(w) > 3 and w not in namen_norm:
                namen_norm.append(w)

        med_entries = []
        for mn in meds:
            form = guess_form(mn, grp_name)
            med_entry = {"name": mn, "form": form}
            med_entries.append(med_entry)
            alle_medikamente.append({
                "name": mn,
                "gruppe_id": grp_id,
                "gruppe_name": grp_name,
                "form": form
            })

        gruppen_list.append({
            "id": grp_id,
            "name": grp_name,
            "namen_normalisiert": namen_norm,
            "medikamente": med_entries,
            "anzahl_medikamente": len(med_entries)
        })

    # Deduplicate alle_medikamente: same name → keep best group (therapeutic over CYP/pharmacokinetic)
    dedup_map = {}  # name_lower -> best entry
    all_groups_map = {}  # name_lower -> list of all gruppe_names
    for med in alle_medikamente:
        key = med['name'].lower()
        if key not in all_groups_map:
            all_groups_map[key] = []
        all_groups_map[key].append(med['gruppe_name'])
        if key not in dedup_map:
            dedup_map[key] = med
        else:
            # Score groups: prefer therapeutic groups over pharmacokinetic categories
            existing_grp = dedup_map[key]['gruppe_name']
            new_grp = med['gruppe_name']
            def group_score(grp, med_name):
                score = 0
                # Penalize CYP/pharmacokinetic/toxicity categories
                if re.search(r'CYP|⛔|⚠|Nephrotox|Hepatotox|QT-|Substrate', grp):
                    score -= 10
                # Bonus if group name is a well-known therapeutic category
                therapeutic = ['Antibiotika', 'Antidiabetika', 'Betablocker', 'ACE-Hemmer',
                    'Statine', 'NSAIDs', 'Schmerzmittel', 'Antihypertensiva', 'Diuretika',
                    'Antikoagulant', 'Corticosteroide', 'Antidepressiva', 'Antiepileptika',
                    'Schilddr', 'Protonenpumpen', 'Antazida', 'Opioide', 'Benzodiazepine']
                for t in therapeutic:
                    if t.lower() in grp.lower():
                        score += 5
                        break
                # Bonus if group contains the medication name
                if med_name.lower() in grp.lower():
                    score += 3
                return score
            if group_score(new_grp, med['name']) > group_score(existing_grp, med['name']):
                dedup_map[key] = med

    # Add alle_gruppen field so the app knows which groups a medication belongs to
    alle_medikamente_dedup = []
    for key, med in sorted(dedup_map.items()):
        med['alle_gruppen'] = all_groups_map.get(key, [med['gruppe_name']])
        alle_medikamente_dedup.append(med)

    alle_medikamente_dedup.sort(key=lambda x: x['name'].lower())
    print(f"  Before dedup: {len(alle_medikamente)} → After dedup: {len(alle_medikamente_dedup)}")

    medikamente_db = {
        "gruppen": gruppen_list,
        "alle_medikamente": alle_medikamente_dedup,
        "meta": {
            "anzahl_gruppen": len(gruppen_list),
            "anzahl_medikamente": len(alle_medikamente_dedup),
            "vor_dedup": len(alle_medikamente)
        }
    }

    print(f"  Normalized groups: {len(gruppen_list)}")
    print(f"  Total medications (unique): {len(alle_medikamente_dedup)}")

    with open(os.path.join(OUTPUT_DIR, 'medikamente_db.json'), 'w', encoding='utf-8') as f:
        json.dump(medikamente_db, f, ensure_ascii=False, indent=2)
    print(f"  Written: medikamente_db.json")

    # Also update medikament_gruppen.json (backwards compatible format)
    gruppen_compat = []
    for grp in gruppen_list:
        gruppen_compat.append({
            "id": grp["id"],
            "name": grp["name"],
            "suchbegriffe": [grp["name"]] + [m["name"] for m in grp["medikamente"][:5]]
        })

    with open(os.path.join(OUTPUT_DIR, 'medikament_gruppen.json'), 'w', encoding='utf-8') as f:
        json.dump(gruppen_compat, f, ensure_ascii=False, indent=2)
    print(f"  Written: medikament_gruppen.json")

    # -------------------------------------------------------
    # 5. KEYWORDS & SYNONYME (keywords_synonyme.json)
    # -------------------------------------------------------
    print("\n=== 5. Keywords & Synonyme ===")
    kw_data = read_sheet(wb, 'Keywords_Synonyme')
    print(f"  Raw rows: {len(kw_data)}")

    keywords = []
    for row in kw_data:
        # Column names from Excel: supplement_id, supplement_name, synonyme_DE, synonyme_EN, schreibweisen, umgangssprache, verwechslungsgefahr, nicht_verwechseln_mit
        kid = safe_str(row.get('supplement_id', row.get('id', '')))
        name = safe_str(row.get('supplement_name', row.get('name', '')))
        if not name:
            continue
        if not kid:
            kid = make_id(name)

        entry = {
            "id": kid,
            "name": name,
            "synonyme": split_list(row.get('synonyme_DE', row.get('synonyme', ''))),
            "synonyme_en": split_list(row.get('synonyme_EN', '')),
            "schreibweisen": split_list(row.get('schreibweisen', '')),
            "umgangssprache": split_list(row.get('umgangssprache', '')),
            "verwechslungsgefahr": split_list(row.get('verwechslungsgefahr', '')),
            "nicht_verwechseln_mit": safe_str(row.get('nicht_verwechseln_mit', '')),
        }

        keywords.append(entry)

    print(f"  Exported: {len(keywords)} keyword entries")

    with open(os.path.join(OUTPUT_DIR, 'keywords_synonyme.json'), 'w', encoding='utf-8') as f:
        json.dump(keywords, f, ensure_ascii=False, indent=2)
    print(f"  Written: keywords_synonyme.json")

    # -------------------------------------------------------
    # 6. FORMS COMPARISON (for reference, optional)
    # -------------------------------------------------------
    print("\n=== 6. Formen Vergleich (optional) ===")
    formen = read_sheet(wb, 'Formen_Vergleich')
    print(f"  Raw rows: {len(formen)}")

    if formen:
        formen_export = []
        for row in formen:
            # Column names: basis_supplement, form_name, form_name_EN, bioverfuegbarkeit_prozent, bioverfuegbarkeit_text, vorteile, nachteile, beste_fuer, preis_niveau, vegan, quelle
            sid = safe_str(row.get('basis_supplement', row.get('supplement_id', row.get('id', ''))))
            form_name = safe_str(row.get('form_name', ''))
            if not sid and not form_name:
                continue

            entry = {
                "supplement_id": sid,
                "form_name": form_name,
                "form_name_en": safe_str(row.get('form_name_EN', '')),
                "bioverfuegbarkeit_prozent": safe_str(row.get('bioverfuegbarkeit_prozent', '')),
                "bioverfuegbarkeit_text": safe_str(row.get('bioverfuegbarkeit_text', '')),
                "vorteile": safe_str(row.get('vorteile', '')),
                "nachteile": safe_str(row.get('nachteile', '')),
                "beste_fuer": safe_str(row.get('beste_fuer', '')),
                "preis_niveau": safe_str(row.get('preis_niveau', '')),
                "vegan": safe_str(row.get('vegan', '')),
                "quelle": safe_str(row.get('quelle', '')),
            }
            formen_export.append(entry)

        with open(os.path.join(OUTPUT_DIR, 'formen_vergleich.json'), 'w', encoding='utf-8') as f:
            json.dump(formen_export, f, ensure_ascii=False, indent=2)
        print(f"  Written: formen_vergleich.json ({len(formen_export)} entries)")

    # -------------------------------------------------------
    # 7. LAENDER REGULIERUNG (laender_regulierung.json)
    # -------------------------------------------------------
    print("\n=== 7. Laender Regulierung ===")
    laender_data = read_sheet(wb, 'Laender_Regulierung')
    print(f"  Raw rows: {len(laender_data)}")

    laender_export = []
    # Country codes from column pattern: XX_max, XX_empf, XX_status, XX_quelle
    COUNTRY_CODES = ['DE','AT','CH','FR','BE','NL','LU','IT','ES','PT','GR','MT','CY',
                     'SE','DK','FI','NO','IS','PL','CZ','SK','HU','RO','BG','SI','HR',
                     'EE','LV','LT','IE','UK']
    for row in laender_data:
        sid = safe_str(row.get('supplement_id', ''))
        name = safe_str(row.get('supplement_name', ''))
        if not sid and not name:
            continue
        if not sid:
            sid = supp_id_map.get(name.lower(), make_id(name))

        entry = {"id": sid, "name": name, "laender": {}}
        for cc in COUNTRY_CODES:
            max_val = safe_str(row.get(f'{cc}_max', ''))
            empf_val = safe_str(row.get(f'{cc}_empf', ''))
            status_val = safe_str(row.get(f'{cc}_status', ''))
            quelle_val = safe_str(row.get(f'{cc}_quelle', ''))
            if max_val or empf_val or status_val:
                entry["laender"][cc] = {
                    "max": max_val,
                    "empfohlen": empf_val,
                    "status": status_val,
                    "quelle": quelle_val
                }
        laender_export.append(entry)

    with open(os.path.join(OUTPUT_DIR, 'laender_regulierung.json'), 'w', encoding='utf-8') as f:
        json.dump(laender_export, f, ensure_ascii=False, indent=2)
    print(f"  Written: laender_regulierung.json ({len(laender_export)} entries)")

    # -------------------------------------------------------
    # 8. DOSIERUNGEN ALTERSGRUPPEN (dosierungen_altersgruppen.json)
    # -------------------------------------------------------
    print("\n=== 8. Dosierungen Altersgruppen ===")
    dos_data = read_sheet(wb, 'Dosierungen_Altersgruppen')
    print(f"  Raw rows: {len(dos_data)}")

    dos_export = []
    for row in dos_data:
        sid = safe_str(row.get('supplement_id', ''))
        name = safe_str(row.get('supplement_name', ''))
        if not sid and not name:
            continue
        if not sid:
            sid = supp_id_map.get(name.lower(), make_id(name))

        entry = {
            "id": sid,
            "name": name,
            "saeugline_0_6m": safe_str(row.get('saeugline_0_6m', '')),
            "saeugline_7_12m": safe_str(row.get('saeugline_7_12m', '')),
            "kinder_1_3j": safe_str(row.get('kinder_1_3j', '')),
            "kinder_4_6j": safe_str(row.get('kinder_4_6j', '')),
            "kinder_7_10j": safe_str(row.get('kinder_7_10j', '')),
            "kinder_11_14j": safe_str(row.get('kinder_11_14j', '')),
            "jugend_15_17j": safe_str(row.get('jugend_15_17j', '')),
            "erwachsene_m": safe_str(row.get('erwachsene_m', '')),
            "erwachsene_w": safe_str(row.get('erwachsene_w', '')),
            "schwangere": safe_str(row.get('schwangere', '')),
            "stillende": safe_str(row.get('stillende', '')),
            "senioren_65plus": safe_str(row.get('senioren_65plus', '')),
            "ul_saeugline": safe_str(row.get('ul_saeugline', '')),
            "ul_kinder": safe_str(row.get('ul_kinder', '')),
            "ul_erwachsene": safe_str(row.get('ul_erwachsene', '')),
            "einheit": safe_str(row.get('einheit', '')),
            "quelle": safe_str(row.get('quelle', '')),
        }
        dos_export.append(entry)

    with open(os.path.join(OUTPUT_DIR, 'dosierungen_altersgruppen.json'), 'w', encoding='utf-8') as f:
        json.dump(dos_export, f, ensure_ascii=False, indent=2)
    print(f"  Written: dosierungen_altersgruppen.json ({len(dos_export)} entries)")

    # -------------------------------------------------------
    # 9. SYMPTOME & ZIELE (symptome_ziele.json)
    # -------------------------------------------------------
    print("\n=== 9. Symptome & Ziele ===")
    sym_data = read_sheet(wb, 'Symptome_Ziele')
    print(f"  Raw rows: {len(sym_data)}")

    sym_export = []
    for row in sym_data:
        sid = safe_str(row.get('id', ''))
        name_de = safe_str(row.get('name_DE', ''))
        if not name_de:
            continue
        if not sid:
            sid = make_id(name_de)

        entry = {
            "id": sid,
            "name_de": name_de,
            "name_en": safe_str(row.get('name_EN', '')),
            "keywords_de": split_list(row.get('keywords_DE', '')),
            "keywords_en": split_list(row.get('keywords_EN', '')),
            "kategorie": safe_str(row.get('kategorie', '')),
            "unterkategorie": safe_str(row.get('unterkategorie', '')),
            "empfohlene_supplements": split_list(row.get('empfohlene_supplements', '')),
            "prioritaet_reihenfolge": split_list(row.get('prioritaet_reihenfolge', '')),
            "ausschluss_symptome": split_list(row.get('ausschluss_symptome', '')),
            "ausschluss_supplements": split_list(row.get('ausschluss_supplements', '')),
            "wichtiger_hinweis_de": safe_str(row.get('wichtiger_hinweis_DE', '')),
            "wichtiger_hinweis_en": safe_str(row.get('wichtiger_hinweis_EN', '')),
            "arzt_pflicht": safe_bool(row.get('arzt_pflicht', False)),
            "arzt_grund": safe_str(row.get('arzt_grund', '')),
        }
        sym_export.append(entry)

    with open(os.path.join(OUTPUT_DIR, 'symptome_ziele.json'), 'w', encoding='utf-8') as f:
        json.dump(sym_export, f, ensure_ascii=False, indent=2)
    print(f"  Written: symptome_ziele.json ({len(sym_export)} entries)")

    # -------------------------------------------------------
    # 10. QUELLEN VERZEICHNIS (quellen_verzeichnis.json)
    # -------------------------------------------------------
    print("\n=== 10. Quellen Verzeichnis ===")
    quellen_data = read_sheet(wb, 'Quellen_Verzeichnis')
    print(f"  Raw rows: {len(quellen_data)}")

    quellen_export = []
    for row in quellen_data:
        sid = safe_str(row.get('supplement_id', ''))
        quelle_name = safe_str(row.get('quelle_name', ''))
        if not sid and not quelle_name:
            continue

        entry = {
            "supplement_id": sid,
            "quelle_typ": safe_str(row.get('quelle_typ', '')),
            "quelle_name": quelle_name,
            "quelle_url": safe_str(row.get('quelle_url', '')),
            "abgerufen_am": safe_str(row.get('abgerufen_am', '')),
            "extrahierte_daten": safe_str(row.get('extrahierte_daten', '')),
            "vertrauenswuerdigkeit": safe_str(row.get('vertrauenswuerdigkeit', '')),
        }
        quellen_export.append(entry)

    with open(os.path.join(OUTPUT_DIR, 'quellen_verzeichnis.json'), 'w', encoding='utf-8') as f:
        json.dump(quellen_export, f, ensure_ascii=False, indent=2)
    print(f"  Written: quellen_verzeichnis.json ({len(quellen_export)} entries)")

    # -------------------------------------------------------
    # 11. STUDIEN EVIDENZ (studien_evidenz.json)
    # -------------------------------------------------------
    print("\n=== 11. Studien & Evidenz ===")
    studien_data = read_sheet(wb, 'Studien_Evidenz')
    print(f"  Raw rows: {len(studien_data)}")

    studien_export = []
    for row in studien_data:
        sid = safe_str(row.get('supplement_id', ''))
        studie_name = safe_str(row.get('studie_name', ''))
        if not sid and not studie_name:
            continue

        entry = {
            "supplement_id": sid,
            "studie_name": studie_name,
            "pmid": safe_str(row.get('pmid', '')),
            "studie_typ": safe_str(row.get('studie_typ', '')),
            "evidenz_level": safe_str(row.get('evidenz_level', '')),
            "teilnehmer_n": safe_str(row.get('teilnehmer_n', '')),
            "indikation": safe_str(row.get('indikation', '')),
            "dosierung_studie": safe_str(row.get('dosierung_studie', '')),
            "ergebnis_kurz": safe_str(row.get('ergebnis_kurz', '')),
            "effektstaerke": safe_str(row.get('effektstaerke', '')),
            "nebenwirkungen": safe_str(row.get('nebenwirkungen', '')),
            "relevanz_fuer_ki": safe_str(row.get('relevanz_fuer_ki', '')),
        }
        studien_export.append(entry)

    with open(os.path.join(OUTPUT_DIR, 'studien_evidenz.json'), 'w', encoding='utf-8') as f:
        json.dump(studien_export, f, ensure_ascii=False, indent=2)
    print(f"  Written: studien_evidenz.json ({len(studien_export)} entries)")

    # -------------------------------------------------------
    # 12. KONTRAINDIKATIONEN DETAIL (kontraindikationen_detail.json)
    # -------------------------------------------------------
    print("\n=== 12. Kontraindikationen Detail ===")
    kontra_data = read_sheet(wb, 'Kontraindikationen_Detail')
    print(f"  Raw rows: {len(kontra_data)}")

    kontra_export = []
    for row in kontra_data:
        sid = safe_str(row.get('supplement_id', ''))
        name = safe_str(row.get('supplement_name', ''))
        kontra = safe_str(row.get('kontraindikation', ''))
        if not sid and not kontra:
            continue
        if not sid:
            sid = supp_id_map.get(name.lower(), make_id(name))

        entry = {
            "supplement_id": sid,
            "supplement_name": name,
            "kontraindikation": kontra,
            "kontraindikation_en": safe_str(row.get('kontraindikation_EN', '')),
            "schweregrad": safe_str(row.get('schweregrad', '')),
            "erklaerung": safe_str(row.get('erklaerung', '')),
            "was_passiert": safe_str(row.get('was_passiert', '')),
            "quelle": safe_str(row.get('quelle', '')),
        }
        kontra_export.append(entry)

    with open(os.path.join(OUTPUT_DIR, 'kontraindikationen_detail.json'), 'w', encoding='utf-8') as f:
        json.dump(kontra_export, f, ensure_ascii=False, indent=2)
    print(f"  Written: kontraindikationen_detail.json ({len(kontra_export)} entries)")

    # -------------------------------------------------------
    # 13. EINHEITEN UMRECHNUNG (einheiten_umrechnung.json)
    # -------------------------------------------------------
    print("\n=== 13. Einheiten Umrechnung ===")
    einh_data = read_sheet(wb, 'Einheiten_Umrechnung')
    print(f"  Raw rows: {len(einh_data)}")

    einh_export = []
    for row in einh_data:
        supplement = safe_str(row.get('supplement', ''))
        if not supplement:
            continue

        entry = {
            "supplement": supplement,
            "supplement_id": supp_id_map.get(supplement.lower(), make_id(supplement)),
            "einheit_1": safe_str(row.get('einheit_1', '')),
            "einheit_2": safe_str(row.get('einheit_2', '')),
            "umrechnung": safe_str(row.get('umrechnung', '')),
            "beispiel": safe_str(row.get('beispiel', '')),
        }
        einh_export.append(entry)

    with open(os.path.join(OUTPUT_DIR, 'einheiten_umrechnung.json'), 'w', encoding='utf-8') as f:
        json.dump(einh_export, f, ensure_ascii=False, indent=2)
    print(f"  Written: einheiten_umrechnung.json ({len(einh_export)} entries)")

    # -------------------------------------------------------
    # 14. KATEGORIEN (kategorien.json)
    # -------------------------------------------------------
    print("\n=== 14. Kategorien ===")
    kat_data = read_sheet(wb, 'Kategorien')
    print(f"  Raw rows: {len(kat_data)}")

    kat_export = []
    for row in kat_data:
        kat_id = safe_str(row.get('kategorie_id', ''))
        kat_name = safe_str(row.get('kategorie_name', ''))
        if not kat_id and not kat_name:
            continue

        entry = {
            "id": kat_id or make_id(kat_name),
            "name": kat_name,
            "name_en": safe_str(row.get('kategorie_name_EN', '')),
            "beschreibung": safe_str(row.get('beschreibung', '')),
        }
        kat_export.append(entry)

    with open(os.path.join(OUTPUT_DIR, 'kategorien.json'), 'w', encoding='utf-8') as f:
        json.dump(kat_export, f, ensure_ascii=False, indent=2)
    print(f"  Written: kategorien.json ({len(kat_export)} entries)")

    # -------------------------------------------------------
    # 15. UEBERSETZUNGEN (uebersetzungen.json)
    # -------------------------------------------------------
    print("\n=== 15. Uebersetzungen ===")
    uebers_data = read_sheet(wb, 'Uebersetzungen')
    print(f"  Raw rows: {len(uebers_data)}")

    uebers_export = []
    for row in uebers_data:
        key = safe_str(row.get('key', ''))
        if not key:
            continue

        entry = {
            "key": key,
            "DE": safe_str(row.get('DE', '')),
            "EN": safe_str(row.get('EN', '')),
            "FR": safe_str(row.get('FR', '')),
            "IT": safe_str(row.get('IT', '')),
            "ES": safe_str(row.get('ES', '')),
            "NL": safe_str(row.get('NL', '')),
            "PL": safe_str(row.get('PL', '')),
        }
        uebers_export.append(entry)

    with open(os.path.join(OUTPUT_DIR, 'uebersetzungen.json'), 'w', encoding='utf-8') as f:
        json.dump(uebers_export, f, ensure_ascii=False, indent=2)
    print(f"  Written: uebersetzungen.json ({len(uebers_export)} entries)")

    # -------------------------------------------------------
    # 16. KI EINSTELLUNGEN (ki_einstellungen.json)
    # -------------------------------------------------------
    print("\n=== 16. KI Einstellungen ===")
    ki_data = read_sheet(wb, 'KI_Einstellungen')
    print(f"  Raw rows: {len(ki_data)}")

    ki_export = []
    for row in ki_data:
        kategorie = safe_str(row.get('KATEGORIE', ''))
        einstellung = safe_str(row.get('EINSTELLUNG', ''))
        if not einstellung:
            continue

        entry = {
            "kategorie": kategorie,
            "einstellung": einstellung,
            "wert_de": safe_str(row.get('WERT_DE', '')),
            "wert_en": safe_str(row.get('WERT_EN', '')),
        }
        ki_export.append(entry)

    with open(os.path.join(OUTPUT_DIR, 'ki_einstellungen.json'), 'w', encoding='utf-8') as f:
        json.dump(ki_export, f, ensure_ascii=False, indent=2)
    print(f"  Written: ki_einstellungen.json ({len(ki_export)} entries)")

    # -------------------------------------------------------
    # 17. LAENDER BEHOERDEN (laender_behoerden.json)
    # -------------------------------------------------------
    print("\n=== 17. Laender Behoerden ===")
    beh_data = read_sheet(wb, 'Laender_Behoerden')
    print(f"  Raw rows: {len(beh_data)}")

    beh_export = []
    for row in beh_data:
        code = safe_str(row.get('Code', ''))
        land = safe_str(row.get('Land', ''))
        if not code and not land:
            continue

        entry = {
            "code": code,
            "land": land,
            "region": safe_str(row.get('Region', '')),
            "behoerde": safe_str(row.get('Behörde', row.get('Behoerde', ''))),
            "website": safe_str(row.get('Website', '')),
            "anmerkung": safe_str(row.get('Anmerkung', '')),
        }
        beh_export.append(entry)

    with open(os.path.join(OUTPUT_DIR, 'laender_behoerden.json'), 'w', encoding='utf-8') as f:
        json.dump(beh_export, f, ensure_ascii=False, indent=2)
    print(f"  Written: laender_behoerden.json ({len(beh_export)} entries)")

    # -------------------------------------------------------
    # 18. EINZELNE GINSENOSIDE (ginsenoside.json) - Spezial
    # -------------------------------------------------------
    print("\n=== 18. Einzelne Ginsenoside ===")
    gin_data = read_sheet(wb, 'Einzelne_Ginsenoside')
    print(f"  Raw rows: {len(gin_data)}")

    gin_export = []
    for row in gin_data:
        gid = safe_str(row.get('ginsenosid_id', ''))
        name = safe_str(row.get('name', ''))
        if not gid and not name:
            continue

        entry = {
            "id": gid or make_id(name),
            "name": name,
            "typ": safe_str(row.get('typ', '')),
            "cas_nummer": safe_str(row.get('cas_nummer', '')),
            "summenformel": safe_str(row.get('summenformel', '')),
            "mw_g_mol": safe_str(row.get('mw_g_mol', '')),
            "pubchem_cid": safe_str(row.get('pubchem_cid', '')),
            "hauptwirkungen": safe_str(row.get('hauptwirkungen', '')),
            "besonderheiten": safe_str(row.get('besonderheiten', '')),
        }
        gin_export.append(entry)

    with open(os.path.join(OUTPUT_DIR, 'ginsenoside.json'), 'w', encoding='utf-8') as f:
        json.dump(gin_export, f, ensure_ascii=False, indent=2)
    print(f"  Written: ginsenoside.json ({len(gin_export)} entries)")

    # -------------------------------------------------------
    # SUMMARY
    # -------------------------------------------------------
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"  wirkstoffe.json:              {len(wirkstoffe)} supplements")
    print(f"  supplement_interactions.json:  {len(supp_interactions)} interactions")
    print(f"  med_supp_interactions.json:    {len(med_supp_interactions)} interactions")
    print(f"  medikamente_db.json:           {len(gruppen_list)} groups, {len(alle_medikamente)} medications")
    print(f"  medikament_gruppen.json:       {len(gruppen_compat)} groups (compat)")
    print(f"  keywords_synonyme.json:        {len(keywords)} entries")
    if formen:
        print(f"  formen_vergleich.json:         {len(formen_export)} entries")
    print(f"  laender_regulierung.json:      {len(laender_export)} entries")
    print(f"  dosierungen_altersgruppen.json:{len(dos_export)} entries")
    print(f"  symptome_ziele.json:           {len(sym_export)} entries")
    print(f"  quellen_verzeichnis.json:      {len(quellen_export)} entries")
    print(f"  studien_evidenz.json:          {len(studien_export)} entries")
    print(f"  kontraindikationen_detail.json:{len(kontra_export)} entries")
    print(f"  einheiten_umrechnung.json:     {len(einh_export)} entries")
    print(f"  kategorien.json:               {len(kat_export)} entries")
    print(f"  uebersetzungen.json:           {len(uebers_export)} entries")
    print(f"  ki_einstellungen.json:         {len(ki_export)} entries")
    print(f"  laender_behoerden.json:        {len(beh_export)} entries")
    print(f"  ginsenoside.json:              {len(gin_export)} entries")
    print(f"\n  Output directory: {OUTPUT_DIR}")
    print("  DONE!")

    wb.close()


if __name__ == '__main__':
    main()
